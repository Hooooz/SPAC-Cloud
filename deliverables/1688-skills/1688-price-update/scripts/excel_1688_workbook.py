from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import List

from openpyxl import load_workbook


HEADER_ROW = 2
FIRST_DATA_ROW = 3
MODEL_COL = 1
COLOR_COL = 2
MOQ_COL = 5
FIRST_PRICE_COL = 6
LINK_HEADER = "链接"


@dataclass
class LadderRow:
    excel_row: int
    quantity_range: str
    current_price: object


@dataclass
class ProductSegment:
    excel_start_row: int
    excel_end_row: int
    model: str
    color: str
    link: str
    ladder_rows: List[LadderRow]
    link_block_start: int
    link_block_end: int
    shared_link_segment_count: int


@dataclass
class LinkBlock:
    start_row: int
    end_row: int
    link: str


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    if not text:
        return ""
    lines = [line.strip() for line in text.splitlines()]
    return "\n".join(line for line in lines if line)


def _normalize_link(value: object) -> str:
    text = _normalize_text(value)
    if "1688.com" not in text:
        return ""
    return text.split("?")[0]


def _normalize_identity(value: object) -> str:
    text = _normalize_text(value)
    if "DISPIMG(" in text.upper():
        return ""
    return text


def _normalize_qty(value: object) -> str:
    text = _normalize_text(value)
    return text.replace("：", "").strip()


def _find_link_col(ws) -> int:
    for col in range(1, ws.max_column + 1):
        if ws.cell(HEADER_ROW, col).value == LINK_HEADER:
            return col
    raise ValueError(f"Could not find `{LINK_HEADER}` in row {HEADER_ROW}")


def _collect_link_blocks(ws, link_col: int) -> List[LinkBlock]:
    blocks: List[LinkBlock] = []
    covered_rows = set()

    merged_ranges = sorted(
        ws.merged_cells.ranges,
        key=lambda cell_range: (cell_range.min_row, cell_range.min_col),
    )
    for cell_range in merged_ranges:
        if cell_range.max_row < FIRST_DATA_ROW:
            continue
        if not (cell_range.min_col <= link_col <= cell_range.max_col):
            continue
        link = _normalize_link(ws.cell(cell_range.min_row, link_col).value)
        if not link:
            continue
        blocks.append(LinkBlock(cell_range.min_row, cell_range.max_row, link))
        covered_rows.update(range(cell_range.min_row, cell_range.max_row + 1))

    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        if row in covered_rows:
            continue
        link = _normalize_link(ws.cell(row, link_col).value)
        if link:
            blocks.append(LinkBlock(row, row, link))

    blocks.sort(key=lambda block: block.start_row)
    return blocks


def extract_product_segments(excel_file: str | Path) -> List[ProductSegment]:
    wb = load_workbook(excel_file, data_only=False)
    ws = wb.active
    link_col = _find_link_col(ws)
    blocks = _collect_link_blocks(ws, link_col)
    segments: List[ProductSegment] = []

    for block in blocks:
        segment_starts = []
        for row in range(block.start_row, block.end_row + 1):
            model = _normalize_text(ws.cell(row, MODEL_COL).value)
            if model:
                segment_starts.append(row)

        if not segment_starts:
            continue

        for index, start_row in enumerate(segment_starts):
            end_row = (
                segment_starts[index + 1] - 1
                if index + 1 < len(segment_starts)
                else block.end_row
            )
            ladder_rows: List[LadderRow] = []
            for row in range(start_row, end_row + 1):
                qty = _normalize_qty(ws.cell(row, MOQ_COL).value)
                if not qty:
                    continue
                ladder_rows.append(
                    LadderRow(
                        excel_row=row,
                        quantity_range=qty,
                        current_price=ws.cell(row, FIRST_PRICE_COL).value,
                    )
                )

            segments.append(
                ProductSegment(
                    excel_start_row=start_row,
                    excel_end_row=end_row,
                    model=_normalize_text(ws.cell(start_row, MODEL_COL).value),
                    color=_normalize_identity(ws.cell(start_row, COLOR_COL).value),
                    link=block.link,
                    ladder_rows=ladder_rows,
                    link_block_start=block.start_row,
                    link_block_end=block.end_row,
                    shared_link_segment_count=len(segment_starts),
                )
            )

    return segments
