"""
更新Excel文件中的价格信息
添加新列：最新价格、价格更新日期
"""
import json
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from excel_1688_workbook import HEADER_ROW

def update_excel_with_prices(excel_file: str, results_dir: str = "price_update_results"):
    """更新Excel文件中的价格信息"""
    
    results_path = Path(results_dir)
    all_results_file = results_path / "all_results.json"
    
    if not all_results_file.exists():
        print(f"错误：找不到结果文件 {all_results_file}")
        return
    
    with open(all_results_file, 'r', encoding='utf-8') as f:
        all_results = json.load(f)
    
    print(f"加载了 {len(all_results)} 条价格记录")
    
    price_map = {}
    for result in all_results:
        excel_row = result['excel_row']
        status = result.get('status', 'success')
        price_range = result.get('price_range', '')
        price_min = result.get('price_min')
        price_max = result.get('price_max')
        
        if status == 'success' and price_min is not None and price_max is not None:
            if price_min == price_max:
                price_text = f"¥{price_min:.2f}"
            else:
                price_text = f"¥{price_min:.2f}-{price_max:.2f}"
        elif status == 'success' and price_range:
            price_text = price_range
        elif status == 'no_price':
            price_text = "无价格"
        elif status in {'invalid', 'error'}:
            price_text = "无效"
        else:
            price_text = "未获取"
        
        price_map[excel_row] = price_text
    
    wb = load_workbook(excel_file)
    ws = wb.active
    
    max_col = ws.max_column
    update_date_col = None
    latest_price_col = None

    for col in range(1, max_col + 1):
        if ws.cell(row=HEADER_ROW, column=col).value == "价格更新日期":
            update_date_col = col
        elif ws.cell(row=HEADER_ROW, column=col).value == "最新1688价格":
            latest_price_col = col

    if update_date_col is None or latest_price_col is None:
        update_date_col = max_col + 1
        latest_price_col = max_col + 2

    ws.cell(row=HEADER_ROW, column=update_date_col, value="价格更新日期")
    ws.cell(row=HEADER_ROW, column=latest_price_col, value="最新1688价格")
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.cell(row=HEADER_ROW, column=update_date_col).font = header_font
    ws.cell(row=HEADER_ROW, column=update_date_col).fill = header_fill
    ws.cell(row=HEADER_ROW, column=latest_price_col).font = header_font
    ws.cell(row=HEADER_ROW, column=latest_price_col).fill = header_fill
    
    update_date = datetime.now().strftime("%Y-%m-%d")
    
    updated_count = 0
    for excel_row, price_text in price_map.items():
        ws.cell(row=excel_row, column=update_date_col, value=update_date)
        ws.cell(row=excel_row, column=latest_price_col, value=price_text)
        
        ws.cell(row=excel_row, column=update_date_col).alignment = Alignment(horizontal='center')
        ws.cell(row=excel_row, column=latest_price_col).alignment = Alignment(horizontal='center')
        
        updated_count += 1
    
    ws.column_dimensions[ws.cell(row=1, column=update_date_col).column_letter].width = 15
    ws.column_dimensions[ws.cell(row=1, column=latest_price_col).column_letter].width = 20
    
    output_file = Path(excel_file).stem + "_价格已更新.xlsx"
    wb.save(output_file)
    
    print(f"\n✓ Excel文件已更新")
    print(f"  更新了 {updated_count} 行价格数据")
    print(f"  新增列：价格更新日期、最新1688价格")
    print(f"  输出文件：{output_file}")
    print(f"  更新日期：{update_date}")
    
    return output_file

if __name__ == '__main__':
    import sys
    
    if len(sys.argv) < 2:
        print("用法: python update_excel.py <excel_file>")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    update_excel_with_prices(excel_file)
