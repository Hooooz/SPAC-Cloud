"""
稳健版阶梯价格更新系统
- 支持断点续传
- 更好的错误处理
- 定期保存进度
"""
import asyncio
import json
import re
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional
from dataclasses import dataclass, asdict
from playwright.async_api import async_playwright, Browser, Page, BrowserContext

from excel_1688_workbook import HEADER_ROW, extract_product_segments


@dataclass
class LadderPrice:
    """阶梯价格"""
    min_qty: int
    max_qty: Optional[int]  # None表示≥
    price: float
    
    def to_dict(self):
        return {
            'min_qty': self.min_qty,
            'max_qty': self.max_qty,
            'price': self.price
        }


PRICE_UNITS = r"(?:个|本|件|套|张|册|包|只|条|台|副|盒|支|对)?"


def _clean_text_lines(body_text: str) -> List[str]:
    return [line.strip() for line in body_text.splitlines() if line.strip()]


def _extract_price_values(text: str) -> List[float]:
    compact = text.replace(" ", "")
    matches = re.findall(r'[¥￥]\s*(\d+(?:\.\d+)?)', compact)
    values = []
    for match in matches:
        value = float(match)
        if 0 < value < 500:
            values.append(value)
    return values


def _extract_joined_price(lines: List[str]) -> Optional[float]:
    compact = ''.join(lines).replace(" ", "")
    match = re.search(r'[¥￥](\d+)(?:\.(\d+))?', compact)
    if not match:
        return None
    value = float(match.group(1) + (f".{match.group(2)}" if match.group(2) else ""))
    if 0 < value < 500:
        return value
    return None


def _find_nearby_price(lines: List[str], index: int) -> Optional[float]:
    windows = [
        lines[max(0, index - 6):index],
        lines[max(0, index - 4):index],
        lines[index + 1:index + 4],
    ]

    for window in windows:
        if not window:
            continue
        value = _extract_joined_price(window)
        if value is not None:
            return value
        values = _extract_price_values(' '.join(window))
        if values:
            return min(values)
    return None


def _parse_page_quantity(line: str) -> Optional[tuple[int, Optional[int], bool]]:
    text = line.replace('：', '').replace(' ', '').strip()

    match = re.match(rf'^(\d+)[-–](\d+){PRICE_UNITS}$', text)
    if match:
        start = int(match.group(1))
        end = int(match.group(2))
        if start <= end:
            return (start, end, False)
        return None

    match = re.match(rf'^≥(\d+){PRICE_UNITS}$', text)
    if match:
        return (int(match.group(1)), None, False)

    match = re.match(rf'^(\d+){PRICE_UNITS}起批$', text)
    if match:
        return (int(match.group(1)), None, True)

    return None


def _normalize_page_prices(prices: List[LadderPrice]) -> List[LadderPrice]:
    deduped: List[LadderPrice] = []
    seen = set()
    for lp in sorted(prices, key=lambda item: (item.min_qty, item.max_qty or 10**9, item.price)):
        key = (lp.min_qty, lp.max_qty, lp.price)
        if key in seen:
            continue
        seen.add(key)
        deduped.append(lp)

    for i, lp in enumerate(deduped[:-1]):
        if lp.max_qty is None:
            next_lp = deduped[i + 1]
            if next_lp.min_qty > lp.min_qty:
                lp.max_qty = next_lp.min_qty - 1

    return deduped


class RobustLadderPriceExtractor:
    """稳健的阶梯价格提取器"""
    
    def __init__(self, cookie_file: str = "1688cookie.json"):
        self.cookie_file = Path(cookie_file)
        self.browser: Optional[Browser] = None
        self.context: Optional[BrowserContext] = None
        self.page: Optional[Page] = None
        self.playwright = None
        
    async def start(self, headless: bool = False):
        """启动浏览器"""
        if not self.playwright:
            self.playwright = await async_playwright().start()
        
        if not self.browser:
            self.browser = await self.playwright.chromium.launch(
                headless=headless,
                args=['--disable-blink-features=AutomationControlled']
            )
        
        if not self.context:
            self.context = await self.browser.new_context(
                user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
                viewport={"width": 1920, "height": 1080}
            )
            
            await self.context.add_init_script("""
                Object.defineProperty(navigator, 'webdriver', {
                    get: () => undefined
                })
            """)
            
            if self.cookie_file.exists():
                cookies = json.loads(self.cookie_file.read_text())
                for cookie in cookies:
                    if 'sameSite' in cookie:
                        if cookie['sameSite'] == 'no_restriction':
                            cookie['sameSite'] = 'None'
                        elif cookie['sameSite'] not in ['Strict', 'Lax', 'None']:
                            cookie['sameSite'] = 'Lax'
                    else:
                        cookie['sameSite'] = 'Lax'
                await self.context.add_cookies(cookies)
                print(f"✓ 已加载 {len(cookies)} 个cookies")
        
        if not self.page:
            self.page = await self.context.new_page()
    
    async def close(self):
        """关闭浏览器"""
        if self.page:
            try:
                await self.page.close()
            except:
                pass
            self.page = None
        
        if self.context:
            try:
                await self.context.close()
            except:
                pass
            self.context = None
        
        if self.browser:
            try:
                await self.browser.close()
            except:
                pass
            self.browser = None
        
        if self.playwright:
            try:
                await self.playwright.stop()
            except:
                pass
            self.playwright = None
    
    async def restart_browser(self, headless: bool = False):
        """重启浏览器"""
        print("  重启浏览器...")
        await self.close()
        await asyncio.sleep(2)
        await self.start(headless)
    
    async def extract_ladder_prices(self, url: str) -> tuple[bool, List[LadderPrice]]:
        """提取阶梯价格"""
        try:
            if not self.page:
                await self.start()
            
            print(f"  访问页面...")
            await self.page.goto(url, timeout=30000, wait_until='domcontentloaded')
            await self.page.wait_for_timeout(5000)
            
            # 检查是否被重定向
            current_url = self.page.url
            if 'offer' not in current_url:
                print(f"  ⚠ 页面被重定向")
                return (False, [])
            
            # 提取价格
            prices = await self._extract_prices()
            
            if prices:
                print(f"  ✓ 提取到 {len(prices)} 个阶梯价格")
                for p in prices:
                    qty_str = f"{p.min_qty}-{p.max_qty}" if p.max_qty else f"≥{p.min_qty}"
                    print(f"    {qty_str}: ¥{p.price}")
                return (True, prices)
            else:
                print(f"  ✗ 未找到阶梯价格")
                return (True, [])
                
        except Exception as e:
            error_msg = str(e)
            if 'Target page, context or browser has been closed' in error_msg:
                print(f"  ⚠ 浏览器已关闭，需要重启")
                return (None, [])  # 需要重启
            elif 'Timeout' in error_msg:
                print(f"  ✗ 超时错误")
                return (False, [])
            else:
                print(f"  ✗ 错误: {error_msg[:100]}")
                return (False, [])
    
    async def _extract_prices(self) -> List[LadderPrice]:
        """从页面提取阶梯价格"""
        try:
            body_text = await self.page.inner_text('body')
            lines = _clean_text_lines(body_text)
            prices: List[LadderPrice] = []

            for index, line in enumerate(lines):
                parsed_qty = _parse_page_quantity(line)
                if not parsed_qty:
                    continue

                min_qty, max_qty, _ = parsed_qty
                price = _find_nearby_price(lines, index)

                if price is not None:
                    prices.append(LadderPrice(min_qty=min_qty, max_qty=max_qty, price=price))

            return _normalize_page_prices(prices)
            
        except Exception as e:
            print(f"  提取错误: {e}")
            return []


def parse_quantity_range(qty_str: str) -> tuple[int, Optional[int]]:
    """解析数量范围"""
    qty_str = re.sub(r'[-–]{2,}', '-', qty_str.replace('：', '').replace(' ', '').strip())
    
    match = re.match(rf'^(\d+)[-–](\d+){PRICE_UNITS}$', qty_str)
    if match:
        return (int(match.group(1)), int(match.group(2)))
    
    match = re.match(rf'^≥(\d+){PRICE_UNITS}$', qty_str)
    if match:
        return (int(match.group(1)), None)
    
    match = re.match(rf'^(\d+){PRICE_UNITS}起批$', qty_str)
    if match:
        max_qty = int(match.group(1)) if int(match.group(1)) > 1 else None
        return (1, max_qty)
    
    match = re.match(rf'^(\d+){PRICE_UNITS}$', qty_str)
    if match:
        return (int(match.group(1)), int(match.group(1)))
    
    return (0, None)


def _normalize_ranges(ranges: List[tuple[int, Optional[int]]]) -> List[tuple[int, Optional[int]]]:
    """将相邻重叠/相接区间收敛为稳定的有效范围。"""
    normalized: List[tuple[int, Optional[int]]] = []
    for index, (min_qty, max_qty) in enumerate(ranges):
        next_min = ranges[index + 1][0] if index + 1 < len(ranges) else None
        effective_max = max_qty

        if next_min is not None and (effective_max is None or effective_max >= next_min):
            effective_max = next_min - 1

        normalized.append((min_qty, effective_max))

    return normalized


def _range_key(min_qty: int, max_qty: Optional[int]) -> tuple[int, int]:
    return (min_qty, max_qty if max_qty is not None else 10**12)


def _same_range(
    left: tuple[int, Optional[int]],
    right: tuple[int, Optional[int]],
) -> bool:
    left_min, left_max = left
    right_min, right_max = right
    if left_min != right_min:
        return False
    if left_max == right_max:
        return True
    if left_max is not None and right_max is not None and abs(left_max - right_max) <= 1:
        return True
    return False


def _range_contains(
    quantity_range: tuple[int, Optional[int]],
    qty: int,
) -> bool:
    min_qty, max_qty = quantity_range
    if qty < min_qty:
        return False
    if max_qty is None:
        return True
    return qty <= max_qty


def _overlap_size(
    left: tuple[int, Optional[int]],
    right: tuple[int, Optional[int]],
) -> int:
    left_min, left_max = left
    right_min, right_max = right
    left_end = left_max if left_max is not None else max(left_min, right_min) + 100000
    right_end = right_max if right_max is not None else max(left_min, right_min) + 100000
    start = max(left_min, right_min)
    end = min(left_end, right_end)
    if end < start:
        return 0
    return end - start + 1


def match_segment_ladder_prices(
    excel_ladders: List[Dict[str, Any]],
    page_prices: List[LadderPrice],
) -> List[Optional[float]]:
    """按整组阶梯匹配，优先保留网页上的真实梯度顺序。"""
    if not excel_ladders or not page_prices:
        return [None for _ in excel_ladders]

    excel_ranges = _normalize_ranges(
        [parse_quantity_range(item['quantity_range']) for item in excel_ladders]
    )
    page_ranges = _normalize_ranges([(item.min_qty, item.max_qty) for item in page_prices])

    results: List[Optional[float]] = [None for _ in excel_ladders]

    # 大多数商品的 Excel 行数与网页阶梯数一致，直接按顺序一一对应。
    if len(excel_ladders) == len(page_prices):
        for index, page_price in enumerate(page_prices):
            results[index] = page_price.price
        return results

    if len(page_prices) == 1:
        page_range = page_ranges[0]
        page_price = page_prices[0].price
        for index, excel_range in enumerate(excel_ranges):
            if _overlap_size(excel_range, page_range) > 0:
                results[index] = page_price
        return results

    page_exact_map = {
        _range_key(min_qty, max_qty): page_price.price
        for (min_qty, max_qty), page_price in zip(page_ranges, page_prices)
    }

    for index, excel_range in enumerate(excel_ranges):
        exact_price = page_exact_map.get(_range_key(*excel_range))
        if exact_price is not None:
            results[index] = exact_price
            continue

        near_matches = []
        for page_index, page_range in enumerate(page_ranges):
            if _same_range(excel_range, page_range):
                near_matches.append((page_index, page_prices[page_index].price))
        if near_matches:
            results[index] = near_matches[0][1]
            continue

        lower_bound_matches = []
        excel_min, excel_max = excel_range
        for page_index, page_range in enumerate(page_ranges):
            if _range_contains(page_range, excel_min):
                order_gap = abs(index - page_index)
                lower_bound_matches.append((-order_gap, page_prices[page_index].price))
        if lower_bound_matches:
            lower_bound_matches.sort(reverse=True)
            results[index] = lower_bound_matches[0][1]
            continue

        candidates = []
        for page_index, page_range in enumerate(page_ranges):
            overlap = _overlap_size(excel_range, page_range)
            if overlap <= 0:
                continue

            page_min, page_max = page_range
            excel_end = excel_max if excel_max is not None else page_min + 100000
            page_end = page_max if page_max is not None else excel_min + 100000
            boundary_gap = abs(excel_min - page_min) + abs(excel_end - page_end)
            order_gap = abs(index - page_index)
            candidates.append((overlap, -boundary_gap, -order_gap, page_prices[page_index].price))

        if candidates:
            candidates.sort(reverse=True)
            results[index] = candidates[0][3]

    return results


def extract_products_from_excel(excel_file: str) -> List[Dict]:
    """从Excel提取产品信息"""
    products = []
    for segment in extract_product_segments(excel_file):
        products.append({
            'excel_start_row': segment.excel_start_row,
            'model': segment.model,
            'color': segment.color,
            'link': segment.link,
            'link_block_start': segment.link_block_start,
            'link_block_end': segment.link_block_end,
            'shared_link_segment_count': segment.shared_link_segment_count,
            'ladder_prices': [
                {
                    'excel_row': ladder_row.excel_row,
                    'quantity_range': ladder_row.quantity_range,
                    'current_price': ladder_row.current_price,
                    'new_price': None,
                }
                for ladder_row in segment.ladder_rows
            ],
        })
    
    return products


def save_progress(products: List[Dict], progress_file: str = "progress.json"):
    """保存进度"""
    with open(progress_file, 'w', encoding='utf-8') as f:
        json.dump({
            'timestamp': datetime.now().isoformat(),
            'products': products
        }, f, ensure_ascii=False, indent=2)


def load_progress(progress_file: str = "progress.json") -> Optional[List[Dict]]:
    """加载进度"""
    if Path(progress_file).exists():
        try:
            with open(progress_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return data['products']
        except:
            pass
    return None


async def update_all_products(
    products: List[Dict],
    cookie_file: str = "1688cookie.json",
    delay: float = 3.0,
    save_interval: int = 10
):
    """更新所有产品的阶梯价格"""
    
    extractor = RobustLadderPriceExtractor(cookie_file)
    await extractor.start()
    link_cache: Dict[str, tuple[Optional[bool], List[LadderPrice]]] = {}
    
    try:
        total = len(products)
        success_count = 0
        fail_count = 0
        
        for i, product in enumerate(products, 1):
            # 跳过已处理的产品
            if product['ladder_prices'] and product['ladder_prices'][0].get('new_price') is not None:
                print(f"\n进度: {i}/{total} - 已处理，跳过")
                continue
            
            print(f"\n{'='*60}")
            print(f"进度: {i}/{total}")
            print(f"产品: {product['model']} - {product['color']}")
            print(f"链接: {product['link']}")
            print(f"{'='*60}")

            if not product['link']:
                for excel_lp in product['ladder_prices']:
                    excel_lp['new_price'] = "无效"
                fail_count += 1
                print("  ✗ 缺少1688链接")
                continue
            
            # 提取阶梯价格
            if product['link'] in link_cache:
                result, page_prices = link_cache[product['link']]
                print("  使用已缓存的链接结果")
            else:
                result, page_prices = await extractor.extract_ladder_prices(product['link'])
                link_cache[product['link']] = (result, page_prices)
            
            # 如果浏览器被关闭，重启
            if result is None:
                await extractor.restart_browser()
                result, page_prices = await extractor.extract_ladder_prices(product['link'])
                link_cache[product['link']] = (result, page_prices)
            
            if result and page_prices:
                matched_prices = match_segment_ladder_prices(product['ladder_prices'], page_prices)
                matched_count = 0
                for excel_lp, matched_price in zip(product['ladder_prices'], matched_prices):
                    excel_qty = excel_lp['quantity_range']

                    if matched_price is not None:
                        excel_lp['new_price'] = matched_price
                        print(f"  ✓ 匹配: {excel_qty} → ¥{matched_price}")
                        matched_count += 1
                    else:
                        print(f"  ✗ 未匹配: {excel_qty}")
                
                if matched_count > 0:
                    success_count += 1
                else:
                    fail_count += 1
                    
            elif result:
                for excel_lp in product['ladder_prices']:
                    excel_lp['new_price'] = "无价格"
                fail_count += 1
                print(f"  ✗ 页面无阶梯价格")
            else:
                for excel_lp in product['ladder_prices']:
                    excel_lp['new_price'] = "无效"
                fail_count += 1
                print(f"  ✗ 链接无效")
            
            # 定期保存进度
            if i % save_interval == 0:
                save_progress(products)
                print(f"\n✓ 已保存进度 ({i}/{total})")
            
            # 延迟
            if i < total:
                await asyncio.sleep(delay)
        
        print(f"\n{'='*60}")
        print(f"统计:")
        print(f"  成功: {success_count}/{total}")
        print(f"  失败: {fail_count}/{total}")
        print(f"{'='*60}")
        
    finally:
        await extractor.close()
    
    return products


def update_excel_file(
    excel_file: str,
    products: List[Dict],
    output_file: str = None
):
    """更新Excel文件"""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    
    wb = load_workbook(excel_file)
    ws = wb.active
    
    max_col = ws.max_column
    today = datetime.now().strftime("%m/%d")
    price_header = f"1688价格\n{today}"
    price_col = None
    date_col = None

    for col in range(1, max_col + 1):
        if ws.cell(row=HEADER_ROW, column=col).value == price_header:
            price_col = col
        elif ws.cell(row=HEADER_ROW, column=col).value == "更新日期":
            date_col = col

    if price_col is None or date_col is None or date_col <= price_col:
        price_col = max_col + 1
        date_col = max_col + 2

    ws.cell(row=HEADER_ROW, column=price_col, value=price_header)
    ws.cell(row=HEADER_ROW, column=date_col, value="更新日期")
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.cell(row=HEADER_ROW, column=price_col).font = header_font
    ws.cell(row=HEADER_ROW, column=price_col).fill = header_fill
    ws.cell(row=HEADER_ROW, column=date_col).font = header_font
    ws.cell(row=HEADER_ROW, column=date_col).fill = header_fill
    
    update_date = datetime.now().strftime("%Y-%m-%d")
    updated_count = 0
    
    for product in products:
        for lp in product['ladder_prices']:
            if lp['new_price'] and lp['new_price'] not in ["无效", "无价格"]:
                excel_row = lp['excel_row']
                ws.cell(row=excel_row, column=price_col, value=lp['new_price'])
                ws.cell(row=excel_row, column=date_col, value=update_date)
                updated_count += 1
            elif lp['new_price']:
                excel_row = lp['excel_row']
                ws.cell(row=excel_row, column=price_col, value=lp['new_price'])
                ws.cell(row=excel_row, column=date_col, value=update_date)
    
    ws.column_dimensions[ws.cell(row=2, column=price_col).column_letter].width = 12
    ws.column_dimensions[ws.cell(row=2, column=date_col).column_letter].width = 12
    
    if not output_file:
        output_file = Path(excel_file).stem + "_阶梯价格已更新.xlsx"
    
    wb.save(output_file)
    
    print(f"\n✓ Excel文件已更新: {output_file}")
    print(f"  更新了 {updated_count} 个阶梯价格")
    
    return output_file


async def main():
    import argparse
    
    parser = argparse.ArgumentParser(description='稳健版阶梯价格更新')
    parser.add_argument('excel_file', help='Excel文件路径')
    parser.add_argument('--cookie-file', default='1688cookie.json', help='Cookie文件')
    parser.add_argument('--delay', type=float, default=3.0, help='请求延迟（秒）')
    parser.add_argument('--output', help='输出文件名')
    parser.add_argument('--resume', action='store_true', help='从上次进度继续')
    
    args = parser.parse_args()
    
    print(f"\n{'='*60}")
    print(f"稳健版阶梯价格更新系统 v3.0")
    print(f"{'='*60}")
    print(f"Excel文件: {args.excel_file}")
    print(f"Cookie文件: {args.cookie_file}")
    print(f"请求延迟: {args.delay}秒")
    print(f"断点续传: {'是' if args.resume else '否'}")
    print(f"{'='*60}\n")
    
    # 加载产品信息
    if args.resume:
        products = load_progress()
        if products:
            print(f"✓ 从进度文件加载 {len(products)} 个产品\n")
        else:
            print("未找到进度文件，从头开始\n")
            products = extract_products_from_excel(args.excel_file)
    else:
        print("正在提取产品信息...")
        products = extract_products_from_excel(args.excel_file)
        print(f"✓ 找到 {len(products)} 个产品\n")
    
    # 更新阶梯价格
    print("开始更新阶梯价格...")
    products = await update_all_products(products, args.cookie_file, args.delay)
    
    # 更新Excel
    update_excel_file(args.excel_file, products, args.output)
    
    # 保存详细结果
    results_file = "阶梯价格更新结果.json"
    with open(results_file, 'w', encoding='utf-8') as f:
        json.dump(products, f, ensure_ascii=False, indent=2)
    print(f"✓ 详细结果: {results_file}")
    
    # 清理进度文件
    if Path("progress.json").exists():
        Path("progress.json").unlink()
        print("✓ 已清理进度文件")


if __name__ == '__main__':
    asyncio.run(main())
